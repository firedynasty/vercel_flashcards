#!/usr/bin/env python3
"""Convert an .xlsx workbook into sheets.json for the flashcards app.

Each sheet becomes a key in the JSON object, with rows as arrays of strings.
Output goes to <name>_xlsx/sheets.json based on the input filename.

Usage:
    python generate_sheets.py xlsxs/spanish.xlsx   -> spanish_xlsx/sheets.json
    python generate_sheets.py xlsxs/chinese.xlsx   -> chinese_xlsx/sheets.json

Deploy:
    rclone copy spanish_xlsx/sheets.json dropbox:/vercel_flashcards/spanish_sheets.json
    rclone link dropbox:/vercel_flashcards/spanish_sheets.json
    # Then update Vercel with the link (append &raw=1):
    vercel env rm DROPBOX_SHEETS_SPANISH_URL production -y
    echo "<LINK>&raw=1" | vercel env add DROPBOX_SHEETS_SPANISH_URL production
"""

import json
import os
import sys

import openpyxl


def main():
    if len(sys.argv) < 2:
        print("Usage: python generate_sheets.py <workbook.xlsx>")
        sys.exit(1)

    path = sys.argv[1]
    # Derive output dir from filename: spanish.xlsx -> spanish_xlsx/
    basename = os.path.splitext(os.path.basename(path))[0]
    out_dir = f"{basename}_xlsx"
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "sheets.json")

    wb = openpyxl.load_workbook(path, read_only=True)
    out = {}

    for name in wb.sheetnames:
        rows = []
        for row in wb[name].iter_rows(values_only=True):
            rows.append([str(c) if c is not None else "" for c in row])
        out[name] = rows

    wb.close()

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False)

    lang = basename.upper()
    print(f"{out_path} written — {len(out)} sheet(s): {', '.join(out.keys())}")
    print()
    print("Deploy commands:")
    print(f"  rclone copy {out_path} dropbox:/vercel_flashcards/{basename}_sheets.json")
    print(f"  rclone link dropbox:/vercel_flashcards/{basename}_sheets.json")
    print(f"  vercel env rm DROPBOX_SHEETS_{lang}_URL production -y")
    print(f'  echo "<LINK>&raw=1" | vercel env add DROPBOX_SHEETS_{lang}_URL production')
    print(f"  # (DROPBOX_SHEETS_URL still works as fallback)")


if __name__ == "__main__":
    main()
